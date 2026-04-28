import csv
from io import BytesIO, StringIO

from openpyxl import load_workbook


def parse_batch_input(
    file_name: str | None,
    file_bytes: bytes | None,
    text: str | None,
    address_column: str = "address",
    max_items: int = 500,
) -> list[str]:
    """
    解析批量输入，支持 CSV / XLSX / 多行文本。

    约束：
    - `file_bytes` 与 `text` 至少提供一个
    - 过滤空行和纯空白
    - 不自动去重，保留原始顺序
    - 超过 `max_items` 时抛出 ValueError
    """
    items: list[str] = []

    if file_name is not None and file_bytes is not None:
        items.extend(_parse_file(file_name, file_bytes, address_column))
    if text and text.strip():
        items.extend(_parse_text_lines(text))

    if not items:
        raise ValueError("file 和 text 至少提供一个")
    if len(items) > max_items:
        raise ValueError(f"单批次最多支持 {max_items} 条地址")
    return items


def _parse_text_lines(text: str) -> list[str]:
    return [line.strip() for line in text.splitlines() if line and line.strip()]


def _parse_file(file_name: str | None, file_bytes: bytes, address_column: str) -> list[str]:
    lowered = (file_name or "").lower()
    if not file_bytes:
        if lowered.endswith(".csv"):
            raise ValueError("CSV 文件为空")
        if lowered.endswith(".xlsx"):
            raise ValueError("Excel 文件为空")
        raise ValueError("上传文件为空")
    if lowered.endswith(".csv"):
        return _parse_csv(file_bytes, address_column)
    if lowered.endswith(".xlsx"):
        return _parse_xlsx(file_bytes, address_column)
    raise ValueError("仅支持 CSV 或 XLSX 文件")


def _parse_csv(file_bytes: bytes, address_column: str) -> list[str]:
    csv_text = _decode_csv_bytes(file_bytes)
    rows = list(csv.reader(StringIO(csv_text)))
    if not rows:
        raise ValueError("CSV 文件为空")

    header = [str(cell).strip() for cell in rows[0]]
    normalized_target = (address_column or "address").strip()
    if normalized_target in header:
        column_index = header.index(normalized_target)
        data_rows = rows[1:]
    else:
        column_index = 0
        data_rows = rows

    values = _extract_column(data_rows, column_index)
    if not values:
        raise ValueError("CSV 文件中未找到有效地址")
    return values


def _parse_xlsx(file_bytes: bytes, address_column: str) -> list[str]:
    workbook = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.worksheets[0] if workbook.worksheets else None
    if sheet is None:
        raise ValueError("Excel 文件为空")

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel 文件为空")

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    normalized_target = (address_column or "address").strip()
    if normalized_target not in header:
        raise ValueError("指定的地址列不存在")

    column_index = header.index(normalized_target)
    values = _extract_column(rows[1:], column_index)
    if not values:
        raise ValueError("Excel 文件中未找到有效地址")
    return values


def _extract_column(rows, column_index: int) -> list[str]:
    items: list[str] = []
    for row in rows:
        if column_index >= len(row):
            continue
        value = row[column_index]
        if value is None:
            continue
        text = str(value).strip()
        if text:
            items.append(text)
    return items


def _decode_csv_bytes(file_bytes: bytes) -> str:
    last_error = None
    for encoding in ("utf-8-sig", "utf-8", "gbk"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError as exc:
            last_error = exc
    raise ValueError(f"CSV 编码无法识别，请使用 UTF-8 或 GBK 编码: {last_error}")
